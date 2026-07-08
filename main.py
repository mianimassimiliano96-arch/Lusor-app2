import os
import json
import datetime
import copy
import zipfile
import tempfile
import shutil
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.camera import Camera
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.filechooser import FileChooserListView
from kivy.clock import Clock
from kivy.utils import platform
from kivy.metrics import dp
from models import Edificio, Stanza, Apparecchio

PIANI = ["Terra", "Primo", "Secondo", "Terzo", "Quarto", "Seminterrato", "Sottotetto", "Altro"]
DESTINAZIONI = ["Ufficio", "Aula/Corsi", "Corridoio", "Magazzino", "Bagno", "Sala riunioni", "Reception", "Cucina", "Altro"]
TIPI_CT = ["Cartongesso", "Grigliato metallico", "Lamellare", "Pannelli fonoassorbenti", "Doghe in legno", "PVC", "Nessuno / A vista", "Altro"]
TIPOLOGIE_APP = ["LED Lineare", "LED Panel", "LED Downlight", "LED Strip", "LED Flood", "Fluorescente", "Alogena", "Incandescenza", "Altro"]
INSTALLAZIONI = ["Soffitto", "Parete", "Pensile", "A sospensione", "A pavimento", "Altro"]
ACCENSIONI = ["Interruttore", "Dimmer", "Sensore presenza", "Crepuscolare", "Timer", "Domotica", "Altro"]

# --- Popup generici ---

class MsgPopup(Popup):
    def __init__(self, titolo, testo, **kwargs):
        super().__init__(title=titolo, size_hint=(0.8, 0.3), **kwargs)
        box = BoxLayout(orientation="vertical", padding=20, spacing=15)
        box.add_widget(Label(text=testo, halign="center"))
        box.add_widget(Button(text="OK", size_hint_y=0.5, on_press=self.dismiss))
        self.add_widget(box)

class ConfirmPopup(Popup):
    def __init__(self, titolo, testo, callback, **kwargs):
        super().__init__(title=titolo, size_hint=(0.8, 0.3), **kwargs)
        self.callback = callback
        box = BoxLayout(orientation="vertical", padding=20, spacing=15)
        box.add_widget(Label(text=testo, halign="center"))
        row = BoxLayout(spacing=10, size_hint_y=0.5)
        row.add_widget(Button(text="Si", on_press=self._yes))
        row.add_widget(Button(text="No", on_press=self.dismiss))
        box.add_widget(row)
        self.add_widget(box)

    def _yes(self, *a):
        self.dismiss()
        if self.callback:
            self.callback()

# --- Home ---

class HomeScreen(Screen):
    def on_enter(self, *a):
        app = App.get_running_app()
        if "nome_input" not in self.ids:
            Clock.schedule_once(lambda x: self.on_enter(), 0.2)
            return
        self.ids.nome_input.text = app.edificio.nome
        self.ids.indirizzo_input.text = app.edificio.indirizzo

    def apri_stanze(self):
        app = App.get_running_app()
        app.edificio.nome = self.ids.nome_input.text.strip()
        app.edificio.indirizzo = self.ids.indirizzo_input.text.strip()
        app.salva_json()
        self.manager.current = "stanze"

    def apri_export(self):
        app = App.get_running_app()
        if not app.edificio.stanze:
            MsgPopup("Attenzione", "Nessuna stanza da esportare.").open()
            return
        self.manager.current = "export"

# --- Lista Stanze ---

class RoomListScreen(Screen):
    def on_enter(self, *a):
        if "lista" not in self.ids:
            Clock.schedule_once(lambda x: self.on_enter(), 0.2)
            return
        self._ricarica()

    def _ricarica(self):
        app = App.get_running_app()
        grid = self.ids.lista
        grid.clear_widgets()
        if not app.edificio.stanze:
            grid.add_widget(Label(text="Nessuna stanza", size_hint_y=None, height=dp(40)))
            return
        for i, s in enumerate(app.edificio.stanze):
            n_foto = len(s.foto_paths)
            n_app = sum(a.quantita for a in s.apparecchi)
            testo = f"{s.id_locale}  |  {s.piano}  |  {s.destinazione}  |  App: {n_app}  |  Foto: {n_foto}"
            card = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(80), padding=dp(8))
            card.add_widget(Label(text=testo, halign="left", size_hint_y=0.5, font_size=dp(13)))
            row = BoxLayout(size_hint_y=0.5)
            b1 = Button(text="Modifica"); b1.bind(on_press=lambda *x, idx=i: self._modifica(idx))
            b2 = Button(text="Duplica"); b2.bind(on_press=lambda *x, idx=i: self._duplica(idx))
            b3 = Button(text="X", background_color=(0.8,0.2,0.2,1)); b3.bind(on_press=lambda *x, idx=i: self._elimina(idx))
            for b in [b1, b2, b3]:
                row.add_widget(b)
            card.add_widget(row)
            grid.add_widget(card)

    def _nuova(self):
        app = App.get_running_app()
        # Replica dati ultima stanza come default
        default = {}
        if app.edificio.stanze:
            ultima = app.edificio.stanze[-1]
            default = {"piano": ultima.piano, "altezza": ultima.altezza,
                       "destinazione": ultima.destinazione, "controsoffitto": ultima.controsoffitto}
        sd = self.manager.get_screen("stanza_detail")
        sd.carica_default(default)
        sd.indice = -1
        self.manager.current = "stanza_detail"

    def _modifica(self, idx):
        sd = self.manager.get_screen("stanza_detail")
        sd.carica_stanza(idx)
        self.manager.current = "stanza_detail"

    def _duplica(self, idx):
        app = App.get_running_app()
        c = copy.deepcopy(app.edificio.stanze[idx])
        c.id_locale += " (copia)"
        c.foto_paths = []
        app.edificio.stanze.append(c)
        app.salva_json()
        self._ricarica()

    def _elimina(self, idx):
        app = App.get_running_app()
        stanza = app.edificio.stanze[idx]
        ConfirmPopup("Elimina", f"Eliminare '{stanza.id_locale}'?", lambda: self._conferma_elimina(idx)).open()

    def _conferma_elimina(self, idx):
        App.get_running_app().edificio.stanze.pop(idx)
        self._ricarica()

# --- Popup Apparecchio ---

class FixturePopup(Popup):
    def __init__(self, apparecchio=None, callback=None, **kwargs):
        super().__init__(title="Apparecchio", size_hint=(0.9, 0.75), **kwargs)
        self.callback = callback
        a = apparecchio or Apparecchio()
        box = BoxLayout(orientation="vertical", padding=10, spacing=5)
        sv = ScrollView()
        form = GridLayout(cols=2, spacing=5, size_hint_y=None, padding=[0,0,0,20])
        form.bind(minimum_height=form.setter("height"))

        form.add_widget(Label(text="Tipologia:"))
        self.tip = Spinner(text=a.tipologia or "LED Panel", values=TIPOLOGIE_APP)
        form.add_widget(self.tip)

        form.add_widget(Label(text="Potenza (W):"))
        self.pot = TextInput(text=a.potenza, input_filter="float")
        form.add_widget(self.pot)

        form.add_widget(Label(text="Installazione:"))
        self.inst = Spinner(text=a.installazione or "Soffitto", values=INSTALLAZIONI)
        form.add_widget(self.inst)

        form.add_widget(Label(text="h_inst (m):"))
        self.hinst = TextInput(text=a.altezza_installazione, input_filter="float")
        form.add_widget(self.hinst)

        form.add_widget(Label(text="Accensione:"))
        self.acc = Spinner(text=a.accensione or "Interruttore", values=ACCENSIONI)
        form.add_widget(self.acc)

        form.add_widget(Label(text="Quantita:"))
        self.qta = TextInput(text=str(a.quantita), input_filter="int")
        form.add_widget(self.qta)

        sv.add_widget(form)
        box.add_widget(sv)
        row = BoxLayout(size_hint_y=0.08, spacing=10)
        row.add_widget(Button(text="Salva", on_press=self._salva))
        row.add_widget(Button(text="Annulla", on_press=self.dismiss))
        box.add_widget(row)
        self.add_widget(box)

    def _salva(self, *a):
        a = Apparecchio(tipologia=self.tip.text, potenza=self.pot.text,
                        installazione=self.inst.text, altezza_installazione=self.hinst.text,
                        accensione=self.acc.text, quantita=int(self.qta.text) if self.qta.text.isdigit() else 1)
        self.dismiss()
        if self.callback:
            self.callback(a)

# --- Dettaglio Stanza ---

class RoomDetailScreen(Screen):
    indice = -1

    def carica_default(self, d):
        self.ids.id_locale.text = ""
        self.ids.piano.text = d.get("piano", "Terra")
        self.ids.altezza.text = d.get("altezza", "3.0")
        self.ids.destinazione.text = d.get("destinazione", "Ufficio")
        self.ids.ct.text = d.get("controsoffitto", "Cartongesso")
        self.apparecchi = []
        self.foto_paths = []
        self.indice = -1
        self._agg_app()
        self._agg_foto()

    def carica_stanza(self, idx):
        app = App.get_running_app()
        s = app.edificio.stanze[idx]
        self.indice = idx
        self.ids.id_locale.text = s.id_locale
        self.ids.piano.text = s.piano if s.piano in PIANI else PIANI[0]
        self.ids.altezza.text = s.altezza
        self.ids.destinazione.text = s.destinazione if s.destinazione in DESTINAZIONI else DESTINAZIONI[0]
        self.ids.ct.text = s.controsoffitto if s.controsoffitto in TIPI_CT else TIPI_CT[0]
        self.apparecchi = copy.deepcopy(s.apparecchi)
        self.foto_paths = list(s.foto_paths)
        self._agg_app()
        self._agg_foto()

    def _agg_app(self):
        g = self.ids.app_list
        g.clear_widgets()
        if not self.apparecchi:
            g.add_widget(Label(text="Nessun apparecchio", size_hint_y=None, height=dp(28)))
            return
        for i, a in enumerate(self.apparecchi):
            t = f"{a.tipologia} | {a.potenza}W | {a.installazione} | x{a.quantita}"
            row = BoxLayout(size_hint_y=None, height=dp(28))
            lbl = Label(text=t, halign="left", size_hint_x=0.7, font_size=dp(11))
            bm = Button(text="M", size_hint_x=0.15, font_size=dp(10))
            bm.bind(on_press=lambda *x, idx=i: self._mod_app(idx))
            bd = Button(text="X", size_hint_x=0.15, font_size=dp(10), background_color=(0.8,0.2,0.2,1))
            bd.bind(on_press=lambda *x, idx=i: self._del_app(idx))
            row.add_widget(lbl); row.add_widget(bm); row.add_widget(bd)
            g.add_widget(row)

    def _agg_foto(self):
        g = self.ids.foto_list
        g.clear_widgets()
        if not self.foto_paths:
            g.add_widget(Label(text="Nessuna foto", size_hint_y=None, height=dp(28)))
            return
        for i, fp in enumerate(self.foto_paths):
            ok = os.path.exists(fp)
            t = os.path.basename(fp) if ok else f"[MISS] {os.path.basename(fp)}"
            row = BoxLayout(size_hint_y=None, height=dp(28))
            row.add_widget(Label(text=t, halign="left", size_hint_x=0.7, font_size=dp(11)))
            bd = Button(text="X", size_hint_x=0.3, font_size=dp(10), background_color=(0.8,0.2,0.2,1))
            bd.bind(on_press=lambda *x, idx=i: self._del_foto(idx))
            row.add_widget(bd)
            g.add_widget(row)

    def _agg_apparecchio(self):
        FixturePopup(callback=lambda a: [self.apparecchi.append(a), self._agg_app()]).open()

    def _mod_app(self, idx):
        FixturePopup(apparecchio=self.apparecchi[idx],
                     callback=lambda a: [self.apparecchi.__setitem__(idx, a), self._agg_app()]).open()

    def _del_app(self, idx):
        self.apparecchi.pop(idx)
        self._agg_app()

    def _scatta_foto(self):
        if platform == "android":
            self._scatta_foto_android()
        else:
            try:
                self._camera_capture(Camera)
            except Exception:
                MsgPopup("Fotocamera", "Fotocamera non disponibile su questo dispositivo.").open()

    def _scatta_foto_android(self):
        try:
            from android.permissions import request_permissions, Permission
            request_permissions([Permission.CAMERA])
        except Exception:
            pass

        app = App.get_running_app()
        os.makedirs(app.foto_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        self._camera_path = os.path.join(app.foto_dir, f"foto_{ts}.jpg")

        try:
            from android import mActivity
            from android.activity import register_activity_result
            from jnius import autoclass

            Intent = autoclass('android.content.Intent')
            MediaStore = autoclass('android.provider.MediaStore')
            ContentValues = autoclass('android.content.ContentValues')
            Uri = autoclass('android.net.Uri')

            values = ContentValues()
            values.put(MediaStore.Images.Media.DISPLAY_NAME, f"foto_{ts}.jpg")
            values.put(MediaStore.Images.Media.MIME_TYPE, "image/jpeg")

            resolver = mActivity.getContentResolver()
            uri = resolver.insert(MediaStore.Images.Media.EXTERNAL_CONTENT_URI, values)
            if uri is None:
                MsgPopup("Fotocamera", "Impossibile creare file.\nUsa 'Sfoglia' per caricare foto.").open()
                return

            intent = Intent(MediaStore.ACTION_IMAGE_CAPTURE)
            intent.putExtra(MediaStore.EXTRA_OUTPUT, uri)
            intent.addFlags(Intent.FLAG_GRANT_WRITE_URI_PERMISSION)

            def on_result(result_code, data):
                if result_code == -1:
                    try:
                        ins = resolver.openInputStream(uri)
                        with open(self._camera_path, 'wb') as f:
                            ba = bytearray(8192)
                            count = ins.read(ba, 0, len(ba))
                            while count > 0:
                                f.write(bytes(ba[:count]))
                                count = ins.read(ba, 0, len(ba))
                        ins.close()
                        resolver.delete(uri, None, None)
                        self._on_foto(self._camera_path)
                    except Exception as e:
                        MsgPopup("Errore", f"Salvataggio foto:\n{e}").open()
                else:
                    MsgPopup("Fotocamera", "Foto annullata.").open()

            register_activity_result(on_result)
            mActivity.startActivityForResult(intent, 100)
        except Exception as e:
            try:
                from kivy.uix.camera import Camera
                self._camera_capture(Camera)
            except Exception as e2:
                MsgPopup("Fotocamera", f"Fotocamera non disponibile.\nUsa 'Sfoglia' per caricare foto.").open()

    def _camera_capture(self, CameraCls):
        app = App.get_running_app()
        os.makedirs(app.foto_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        self._camera_path = os.path.join(app.foto_dir, f"foto_{ts}.jpg")

        cbox = BoxLayout(orientation="vertical", spacing=5, padding=5)
        cam = CameraCls(play=True, resolution=(640, 480))
        cbox.add_widget(cam)
        btn_box = BoxLayout(size_hint_y=None, height=dp(44), spacing=10)
        btn_box.add_widget(Button(text="Scatta", on_press=lambda *x: self._camera_salva(cam, cbox)))
        btn_box.add_widget(Button(text="Annulla", on_press=lambda *x: [setattr(cam, 'play', False), cbox.parent.parent.dismiss()]))
        cbox.add_widget(btn_box)

        popup = Popup(title="Fotocamera", content=cbox, size_hint=(0.9, 0.7))
        popup.open()

    def _camera_salva(self, cam, cbox):
        try:
            cam.export_to_png(self._camera_path)
        except Exception:
            from kivy.core.image import Image as CoreImage
            from kivy.graphics import Color, Rectangle
            from io import BytesIO
            try:
                texture = cam.texture
                if texture:
                    core_img = CoreImage(texture)
                    core_img.save(self._camera_path, flipped=False)
                else:
                    raise Exception("Nessun frame dalla fotocamera")
            except Exception as e2:
                MsgPopup("Errore", f"Impossibile salvare foto:\n{e2}").open()
                return
        cam.play = False
        cbox.parent.parent.dismiss()
        self._on_foto(self._camera_path)

    def _aggiungi_foto(self):
        try:
            if platform == "android":
                self._apri_galleria_android()
            else:
                from plyer import filechooser
                filechooser.open_file(on_selection=lambda s: self._on_foto(s[0] if s else None))
        except Exception:
            self._fc_fallback()

    def _apri_galleria_android(self):
        try:
            from android import mActivity
            from android.activity import register_activity_result
            from jnius import autoclass

            Intent = autoclass('android.content.Intent')
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType("image/*")
            intent.addCategory(Intent.CATEGORY_OPENABLE)

            app = App.get_running_app()
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            self._gallery_output_path = os.path.join(app.foto_dir, f"foto_{ts}.jpg")

            def on_result(result_code, data):
                if result_code == -1 and data:
                    uri = data.getData()
                    if uri:
                        try:
                            resolver = mActivity.getContentResolver()
                            ins = resolver.openInputStream(uri)
                            with open(self._gallery_output_path, 'wb') as f:
                                ba = bytearray(8192)
                                count = ins.read(ba, 0, len(ba))
                                while count > 0:
                                    f.write(bytes(ba[:count]))
                                    count = ins.read(ba, 0, len(ba))
                            ins.close()
                            self._on_foto(self._gallery_output_path)
                        except Exception as e:
                            MsgPopup("Errore", f"Impossibile aprire immagine:\n{e}").open()
                elif result_code != -1:
                    MsgPopup("Galleria", "Nessuna immagine selezionata.").open()

            register_activity_result(on_result)
            mActivity.startActivityForResult(intent, 200)
        except Exception as e:
            self._fc_fallback()

    def _fc_fallback(self):
        c = BoxLayout(orientation="vertical", padding=10)
        fc = FileChooserListView(filters=["*.jpg", "*.jpeg", "*.png", "*.bmp"])
        p = Popup(title="Seleziona foto", content=c, size_hint=(0.9, 0.8))
        br = BoxLayout(size_hint_y=0.12, spacing=10)
        br.add_widget(Button(text="Seleziona", on_press=lambda *x: [p.dismiss(), self._on_foto(fc.selection[0] if fc.selection else None)]))
        br.add_widget(Button(text="Annulla", on_press=p.dismiss))
        c.add_widget(fc); c.add_widget(br)
        p.open()

    def _on_foto(self, path):
        if path and os.path.exists(path):
            self.foto_paths.append(path)
            self._agg_foto()

    def _del_foto(self, idx):
        self.foto_paths.pop(idx)
        self._agg_foto()

    def _salva(self):
        idl = self.ids.id_locale.text.strip()
        if not idl:
            MsgPopup("Attenzione", "Inserisci l'ID del locale.").open()
            return
        s = Stanza(id_locale=idl, piano=self.ids.piano.text, altezza=self.ids.altezza.text,
                   destinazione=self.ids.destinazione.text, controsoffitto=self.ids.ct.text,
                   apparecchi=self.apparecchi,
                   foto_paths=[p for p in self.foto_paths if os.path.exists(p)])
        app = App.get_running_app()
        if 0 <= self.indice < len(app.edificio.stanze):
            app.edificio.stanze[self.indice] = s
        else:
            app.edificio.stanze.append(s)
        app.salva_json()
        self.manager.current = "stanze"

# --- Export ---

class ExportScreen(Screen):
    def on_enter(self, *a):
        if "info" not in self.ids:
            Clock.schedule_once(lambda x: self.on_enter(), 0.2)
            return
        app = App.get_running_app()
        ns = len(app.edificio.stanze)
        na = sum(sum(x.quantita for x in s.apparecchi) for s in app.edificio.stanze)
        nf = sum(len(s.foto_paths) for s in app.edificio.stanze)
        self.ids.info.text = f"Edificio: {app.edificio.nome}\nStanze: {ns}\nApparecchi: {na}\nFoto: {nf}"

    def _esporta_excel(self):
        app = App.get_running_app()
        if not app.edificio.stanze:
            MsgPopup("Attenzione", "Nessuna stanza.").open()
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Riepilogo"
            ws.cell(row=1, column=1, value="RAPPORTO DI RILIEVO - RELAMPING").font = Font(bold=True, size=14)
            ws.cell(row=3, column=1, value="Edificio:").font = Font(bold=True)
            ws.cell(row=3, column=2, value=app.edificio.nome)
            ws.cell(row=4, column=1, value="Indirizzo:").font = Font(bold=True)
            ws.cell(row=4, column=2, value=app.edificio.indirizzo)
            tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            ofill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            hdrs = ["ID Locale", "Piano", "Altezza (m)", "Destinazione", "Controsoffitto", "N. App.", "Foto"]
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=7, column=ci, value=h)
                c.font = hf; c.fill = hfill; c.border = tb; c.alignment = Alignment(horizontal="center")
            for i, s in enumerate(app.edificio.stanze):
                r = i + 8
                v = [s.id_locale, s.piano, s.altezza, s.destinazione, s.controsoffitto,
                     sum(a.quantita for a in s.apparecchi),
                     str(len(s.foto_paths)) if s.foto_paths else "No"]
                for col, val in enumerate(v, 1):
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.border = tb
                    if i % 2: cell.fill = ofill
            ws2 = wb.create_sheet("Dettaglio Apparecchi")
            hdrs2 = ["ID Locale", "Tipologia", "Potenza (W)", "Installazione", "h_inst (m)", "Accensione", "Quantita"]
            for ci, h in enumerate(hdrs2, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.font = hf; c.fill = hfill; c.border = tb
            ri = 2
            for s in app.edificio.stanze:
                for a in s.apparecchi:
                    for col, v in enumerate([s.id_locale, a.tipologia, a.potenza, a.installazione,
                                             a.altezza_installazione, a.accensione, a.quantita], 1):
                        cell = ws2.cell(row=ri, column=col, value=v)
                        cell.border = tb
                    ri += 1
            ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 12; ws.column_dimensions["D"].width = 22
            ws.column_dimensions["E"].width = 22; ws.column_dimensions["F"].width = 10
            ws.column_dimensions["G"].width = 8
            for col in "ABCDEFG":
                ws2.column_dimensions[col].width = 16

            percorso = os.path.join(app.foto_dir, f"rilievo_{app.edificio.nome.replace(' ', '_')}.xlsx")
            wb.save(percorso)

            if any(s.foto_paths for s in app.edificio.stanze):
                zip_path = percorso.rsplit(".",1)[0] + "_foto.zip"
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for s in app.edificio.stanze:
                        if s.foto_paths:
                            d = "foto/" + s.id_locale.replace("/","_").replace("\\","_") + "/"
                            for fp in s.foto_paths:
                                if os.path.exists(fp):
                                    zf.write(fp, d + os.path.basename(fp))
                self._condividi_file(zip_path)
            self._condividi_file(percorso)
        except ImportError:
            MsgPopup("Errore", "openpyxl non installato.").open()
        except Exception as e:
            MsgPopup("Errore", str(e)).open()

    def _esporta_lsr(self):
        app = App.get_running_app()
        percorso = os.path.join(app.foto_dir, f"rilievo_{app.edificio.nome.replace(' ', '_')}.lsr")
        try:
            data = app.edificio.to_dict()
            with tempfile.TemporaryDirectory() as tmp:
                with open(os.path.join(tmp, "rilievo.json"), "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                fd = os.path.join(tmp, "foto")
                os.makedirs(fd, exist_ok=True)
                for s in app.edificio.stanze:
                    for idx, fp in enumerate(s.foto_paths):
                        if os.path.exists(fp):
                            b = os.path.basename(fp)
                            n, e = os.path.splitext(b)
                            shutil.copy2(fp, os.path.join(fd, f"{n}_{idx}{e}"))
                with zipfile.ZipFile(percorso, "w", zipfile.ZIP_DEFLATED) as zf:
                    for root, _, files in os.walk(tmp):
                        for fn in files:
                            zf.write(os.path.join(root, fn), os.path.relpath(os.path.join(root, fn), tmp))
            self._condividi_file(percorso)
            MsgPopup("Salvato", f"Progetto salvato:\n{percorso}").open()
        except Exception as e:
            MsgPopup("Errore", str(e)).open()

    def _condividi_file(self, path):
        if not os.path.exists(path):
            return
        try:
            if platform == "android":
                from android import mActivity
                from jnius import autoclass
                Intent = autoclass('android.content.Intent')
                Uri = autoclass('android.net.Uri')
                File = autoclass('java.io.File')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                intent = Intent(Intent.ACTION_SEND)
                intent.setType("*/*")
                file = File(path)
                uri = Uri.fromFile(file)
                intent.putExtra(Intent.EXTRA_STREAM, uri)
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                currentActivity = PythonActivity.mActivity
                currentActivity.startActivity(Intent.createChooser(intent, "Condividi"))
            else:
                from plyer import filechooser
                # Su desktop salviamo già in percorso fisso, mostriamo solo messaggio
                pass
        except Exception:
            pass

# --- App principale ---

class LusorApp(App):
    def build(self):
        self.icon = os.path.join(os.path.dirname(__file__), "icon.png")
        self.edificio = Edificio()
        self.data_dir = self._get_data_dir()
        os.makedirs(self.data_dir, exist_ok=True)
        self.foto_dir = os.path.join(self.data_dir, "foto")
        os.makedirs(self.foto_dir, exist_ok=True)
        sm = ScreenManager()
        sm.add_widget(HomeScreen(name="home"))
        sm.add_widget(RoomListScreen(name="stanze"))
        sm.add_widget(RoomDetailScreen(name="stanza_detail"))
        sm.add_widget(ExportScreen(name="export"))
        Clock.schedule_once(lambda x: sm.get_screen("home").on_enter(), 0.3)
        return sm

    def _get_data_dir(self):
        if platform == "android":
            from android.storage import app_storage_path
            return app_storage_path()
        return os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "Lusor")

    def salva_json(self):
        try:
            path = os.path.join(self.data_dir, "rilievo.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.edificio.to_dict(), f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Errore salvataggio: {e}")

    def mostra_messaggio(self, titolo, testo):
        MsgPopup(titolo, testo).open()

if __name__ == "__main__":
    LusorApp().run()
